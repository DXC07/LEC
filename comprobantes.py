from functions import Functions
import os
import pandas as pd
from datetime import datetime
import shutil
import openpyxl as xl

class Comprobantes(Functions):

    def __init__(self, date):
        super().__init__()
        self.date = date # Fecha del último día de los rechazos gestionados
        self.df_cons = self.read_consolidation_file() # Consolidado rechazos
        self.df_trns_agiles_gn, self.df_trns_agiles_bv = self._get_trns_agiles_file() # Archivo paramétrico de trns ágiles generico

    def _get_trns_agiles_file(self):
        """Carga archivo paramétrico de transacciones ágiles"""
        # Abrir archivo comprobantes genéricos
        df_gn = pd.read_excel(os.path.join(self.static, self.settings['comprobantes_contables']['trns_agiles']), header = 1, sheet_name = 'Genérico')

        # Abrir archivo comprobantes BV
        df_bv = pd.read_excel(os.path.join(self.static, self.settings['comprobantes_contables']['trns_agiles']), header = 1, sheet_name = 'BV')

        # Llenar vacíos de celdas combinadas
        df_bv['CTA CONT'].fillna(method = 'ffill', inplace = True)
        df_bv['CLASIFICACIÓN'].fillna(method = 'ffill', inplace = True)
        df_bv['CLASIFICACIÓN O TRN VALIDAR'].fillna(method = 'ffill', inplace = True)

        # Expandir trns separadas por coma
        df_bv['CLASIFICACIÓN O TRN VALIDAR'] = df_bv['CLASIFICACIÓN O TRN VALIDAR'].str.split(',')
        df_bv = df_bv.explode('CLASIFICACIÓN O TRN VALIDAR').reset_index(drop = True)
        df_bv['CLASIFICACIÓN O TRN VALIDAR'] = df_bv['CLASIFICACIÓN O TRN VALIDAR'].str.strip()

        # Crear llave para hacer cruce
        def func(row):
            if row['CTA CONT'] != 'No aplica' and row['CLASIFICACIÓN O TRN VALIDAR'] == 'No aplica':
                return str(int(row['CTA CONT']))
            elif row['CLASIFICACIÓN O TRN VALIDAR'] != 'No aplica':
                return row['CLASIFICACIÓN O TRN VALIDAR']
            else:
                return row['CLASIFICACIÓN']

        df_bv['llave_trns_agiles'] = df_bv.apply(func, axis = 1)

        df_bv.to_excel('./out/trnagilesbv.xlsx', index = False)

        return df_gn, df_bv
    
    def read_consolidation_file(self):
        """Carga la información que se encuentra en el consolidado del día"""
        # Definir ruta archivo rechazos
        self.rej_folder = os.path.join(self.paths['carpeta_rechazos'], self.date.strftime('%Y'), str(self.date.month) + '. ' + self.date.strftime('%B').upper(), self.settings['consolidado_rechazos']['nombre_carpeta'])
        self.consolidation_file = os.path.join(self.rej_folder, self.settings['consolidado_rechazos']['nombre'].replace('AAAAMMDD', self.date.strftime('%Y%m%d')))

        # Cargar hoja de rechazos en Dataframe
        df = pd.read_excel(self.consolidation_file, sheet_name = self.settings['consolidado_rechazos']['hoja_rechazos'])

        # Poner positivo el monto
        df['Monto'] = df.apply(lambda row: abs(row['Monto']), axis = 1)

        return df
    
    def gen_comp_gn_file(self):
        """Genera y guarda el comprobante contable genérico"""

        # Definir rechazos que se van a incluir en el comprobante contable de acuerdo a la clasificación y a las transacciones
        classification = ['PAGOS', 'RECAUDOS', 'CXC', 'CXP']
        trns = ['S9014', 'S9015', 'D9193', 'D9194']

        # Filtrar consolidado para generar comprobante contable genérico
        df_gn = self.df_cons[(self.df_cons['Clasificación'].isin(classification)) | (self.df_cons['Concatenar'].isin(trns))].reset_index(drop = True)

        if df_gn.shape[0] > 0:
            # Crear campo llave para realizar cruce
            df_gn['llave_trns_agiles'] = df_gn.apply(lambda row: row['Clasificación'] if row['Clasificación'] in classification else row['Concatenar'], axis = 1)

            # Hacer merge con archivo transacciones ágiles
            df_gn = pd.merge(df_gn, self.df_trns_agiles_gn, how = 'left', left_on = 'llave_trns_agiles', right_on = 'CLASIFICACIÓN CONSOLIDADO', validate = 'm:1')

            # Llenar comprobante
            df_comp_gn = pd.DataFrame()
            df_comp_gn['CÓDIGO'] = df_gn['CÓDIGO DE TRANSACCIÓN']
            df_comp_gn['DESCRIPCIÓN'] = df_gn['NOMBRE DE LA TRANSACCIÓN']
            df_comp_gn['OFICINA DÉBITO'] = df_gn['Sucursal libros']
            df_comp_gn['DIA PROCESO'] = datetime.now().day
            df_comp_gn['MES PROCESO'] = datetime.now().month
            df_comp_gn['AÑO PROCESO'] = int(datetime.now().strftime('%y'))
            df_comp_gn['DIA CONTABILIZACIÓN'] = datetime.now().day
            df_comp_gn['MES CONTABILIZACIÓN'] = datetime.now().month
            df_comp_gn['AÑO CONTABILIZACIÓN'] = int(datetime.now().strftime('%y'))
            df_comp_gn['OFICINA CRÉDITO'] = df_gn['Sucursal libros']
            df_comp_gn['NUM COMPROBANTE'] = 770200
            df_comp_gn['TRN'] = 88
            df_comp_gn['TERCERO'] = df_gn.apply(lambda row: int(row['Número identificación']) if row['CÓDIGO DE TRANSACCIÓN'] in ["DE0003", "DE0005", "DE0007", "DE0040"] else 0, axis = 1)
            df_comp_gn['CAMPO C'] = df_gn.apply(lambda row: int(row['Fecha efectiva']) if row['CÓDIGO DE TRANSACCIÓN'] in ["DE0003", "DE0005", "DE0007", "DE0040"] else 0, axis = 1)
            df_comp_gn['TERCERO OFICINA CRÉDITO'] = df_gn.apply(lambda row: 0 if row['CÓDIGO DE TRANSACCIÓN'] in ["DE0003", "DE0005", "DE0007", "DE0040"] else int(row['Número identificación']), axis = 1)
            df_comp_gn['CAMPO C OFICINA CRÉDITO'] = df_gn.apply(lambda row: 0 if row['CÓDIGO DE TRANSACCIÓN'] in ["DE0003", "DE0005", "DE0007", "DE0040"] else int(row['Fecha efectiva']), axis = 1)

            # Función para llenar campo detalle
            def func(row):
                if row['Clasificación'] in ['PAGOS', 'RECAUDOS']:
                    return 'Se contabiliza rechazo de ' + row['Clasificación'].lower()
                elif row['Clasificación'] in ['CXC', 'CXP']:
                    return 'Se contabiliza rechazo a ' + row['Clasificación'].lower()
                else:
                    return row['Clasificación']

            df_comp_gn['DETALLE1'] = df_gn.apply(func, axis = 1)
            df_comp_gn['CUENTA DÉBITO 1'] = df_gn['Monto']
            df_comp_gn['CUENTA CRÉDITO 1'] = df_gn['Monto']

            # Exportar a excel archivo comprobantes
            df_comp_gn.to_excel('./out/comprobantes_gn.xlsx', index = False)

            # Definir ruta comprobante
            comp_folder = os.path.join(self.paths['carpeta_rechazos'], self.date.strftime('%Y'), str(self.date.month) + '. ' + self.date.strftime('%B').upper(), self.settings['comprobantes_contables']['nombre_carpeta'])
            comp_file = os.path.join(comp_folder, self.settings['comprobantes_contables']['nombre_gn'].replace('AAAAMMDD', datetime.now().strftime('%Y%m%d')))

            # Crear carpeta
            if not os.path.exists(comp_folder):
                os.makedirs(comp_folder)
            
            # Copiar planilla en la ruta
            shutil.copy(os.path.join(self.static, 'PlantillaTrxAgiles GN.xlsx'), comp_file)

            # Formatos comprobante
            cells_format = {
                'CUENTA DÉBITO 1': {
                    'number_format': '0.00'
                },
                'CUENTA CRÉDITO 1': {
                    'number_format': '0.00'
                }
            }

            # Llenar archivo
            self.update_excel_file(comp_file, 'GENERICO', df_comp_gn, cells_format = cells_format)

            # Guardar soporte
            wb = xl.load_workbook(comp_file)
            sheet = wb['soporte']

            for j, column in enumerate(df_gn.columns, 1):
                cell = sheet.cell(row = 1, column = j)
                cell.value = column
            
            self.fill_excel_sheet(sheet, df_gn)

            wb.save(comp_file)
        else:
            msg = 'No se encontraron rechazos para realizar comprobante contable genérico'
            self.show_messages(msg, 'warning', True)
    
        return None

    def gen_comp_bv_file(self):
        """Genera y guarda el comprobante contable genérico"""
        # Definir rechazos que se van a incluir en el comprobante contable de acuerdo a la clasificación, las transacciones la cuenta contable
        classification = ['MEDIOS DE PAGO', 'FINANCIACION', 'TRN PARA CONTABILIZAR', 'LECTURA BATCH', 'EMBARGOS LECTURA TIEMPO REAL']

        # Filtrar consolidado para generar comprobante contable genérico
        df_bv = self.df_cons[((self.df_cons['Clasificación'].isin(classification)) | ((self.df_cons['Código TRN rechazo'].astype(int) == 0)) & (self.df_cons['Clasificación'] == 'LECTURA TIEMPO REAL')) & (self.df_cons['Estado rechazo'] != 'NO APLICADO')]

        if df_bv.shape[0] > 0:
            # Crear campo llave para realizar cruce
            def func(row):
                if row['Concatenar'] in self.df_trns_agiles_bv['CLASIFICACIÓN O TRN VALIDAR'].to_list():
                    return row['Concatenar']
                elif row['Clasificación'] in ['MEDIOS DE PAGO', 'FINANCIACION']:
                    return str(int(row['CTA CONTABLE']))
                elif int(row['Código TRN rechazo']) == 0 and row['Concatenar'] not in self.df_trns_agiles_bv['CLASIFICACIÓN O TRN VALIDAR'].to_list():
                    return 'SIN TRN RECHAZO'
                elif row['Clasificación'] in ['LECTURA BATCH', 'EMBARGOS LECTURA TIEMPO REAL']:
                    return row['Clasificación']

            df_bv['llave_trns_agiles'] = df_bv.apply(func, axis = 1)

            # Crear dataframe soporte
            df_bv_sup = df_bv.reset_index(drop = True).copy()

            # Hacer merge con archivo transacciones ágiles
            df_bv = pd.merge(df_bv, self.df_trns_agiles_bv, how = 'left', on = 'llave_trns_agiles')

            # Validar los códigos que no hacen match
            if df_bv['CODIGO DE TRANSACCIÓN'].isnull().sum() > 0:
                self.show_messages('Hay registros sin código de transacción ágil', 'warning', True)

            df_bv.to_excel('./out/merge_bv.xlsx', index = False)

            # Llenar comprobante
            df_comp_bv = pd.DataFrame()
            df_comp_bv['Código Transacción'] = df_bv['CODIGO DE TRANSACCIÓN']
            df_comp_bv['Descripción'] = df_bv['NOMBRE DE LA TRANSACCIÓN']
            df_comp_bv['Comentario'] = df_bv.apply(lambda row: 'Rechazo contabilizado al área ' + row['Clasificación'] if row['Clasificación'] in [c for c in classification if c != 'VALIDAR'] else row['Descripción TRN'], axis = 1)
            df_comp_bv['Fecha Efectiva'] = int(datetime.now().strftime('%Y%m%d'))
            df_comp_bv['Monto'] = df_bv['Monto']
            df_comp_bv['Sucursal Contable'] = df_bv.apply(lambda row: row['OFICINA CONTABILIDAD'] if str(row['OFICINA CONTABILIDAD']) != 'nan' else row['Sucursal libros'], axis = 1)
            df_comp_bv['Sucursal Origen'] = df_comp_bv['Sucursal Contable']

            # Exportar a excel archivo comprobantes
            df_comp_bv.to_excel('./out/comprobantes_bv.xlsx', index = False)

            # Definir ruta comprobante
            comp_folder = os.path.join(self.paths['carpeta_rechazos'], self.date.strftime('%Y'), str(self.date.month) + '. ' + self.date.strftime('%B').upper(), self.settings['comprobantes_contables']['nombre_carpeta'])
            comp_file = os.path.join(comp_folder, self.settings['comprobantes_contables']['nombre_bv'].replace('AAAAMMDD', datetime.now().strftime('%Y%m%d')))

            # Crear carpeta
            if not os.path.exists(comp_folder):
                os.makedirs(comp_folder)
            
            # Copiar planilla en la ruta
            shutil.copy(os.path.join(self.static, 'PlantillaTrxAgiles BV.xlsx'), comp_file)

            # Llenar archivo
            self.update_excel_file(comp_file, 'Transacciones', df_comp_bv, clean_sheet = False)

            # Guardar soporte
            wb = xl.load_workbook(comp_file)
            sheet = wb['soporte']

            for j, column in enumerate(df_bv_sup.columns, 1):
                cell = sheet.cell(row = 1, column = j)
                cell.value = column
            
            self.fill_excel_sheet(sheet, df_bv_sup)

            wb.save(comp_file)
        else:
            msg = 'No se encontraron rechazos para realizar comprobante contable BANKVISION'
            self.show_messages(msg, 'warning', True)
    

    def run(self):
        """Orquesta la ejecución de la clase"""
        self.gen_comp_gn_file()
        self.gen_comp_bv_file()
        
        return None