from functions import Functions
import os
import pandas as pd
from datetime import datetime, timedelta
import shutil
import openpyxl as xl
from copy import copy
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException, NoSuchElementException, ElementNotInteractableException, StaleElementReferenceException
import json
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from time import sleep
import re
import win32com.client as win32com
from dateutil.relativedelta import relativedelta

class Lectura(Functions):

    def __init__(self, credentials, date):
        super().__init__()
        self.red_user = credentials['login_RED_user'] + '@bancolombia.com.co' # usuario para autenticarse en sitio de lectura
        self.red_pwd = credentials['login_RED_pwd'] # contraseña para autenticarse en sitio de lectura
        self.date = date # Fecha del último día de los rechazos gestionados
        self.consolidation_file = self._get_consolidation_file() # Ruta del consolidado de rechazos que se está procesando
        self.lectura_path = self._get_lectura_path() # Ruta de la lectura que se está procesando
        self.driver = None # Driver para interactuar en el sitio de lecturas
        self.wait = None # Espera dinámica driver
        self.app_list = self._get_app_list() # Lista de aprobadores de lectura
        self.rad_number = self._get_rad_number() # Número de radicado de la lectura
        self.rad_state = None # Estado de la solictud de lectura
        self.lectura_reply_path = None # Ruta de la respuesta de la lectura
        self.df_soporte = None # Soporte que se envía en la lectura

        # Abrir archivo xpaths
        with open(os.path.join(self.static, 'xpaths.json'), encoding='utf-8') as f:
            self.xpaths = json.load(f)
        
    def _get_rad_number(self):
        """Obtiene el número de radicado de la lectura en caso de que ya se haya procesado"""
        # Abrir archivo procesamiento
        with open('./procesamiento.json', mode = 'r', encoding = 'utf8') as jsonfile:
            data = json.load(jsonfile)

        try:
            rad_number = data['radicado_lectura']

            return rad_number
        except KeyError:
            
            return None
    
    def _get_consolidation_file(self):
        """Obtiene la ruta del consolidado de rechazos que se está procesando"""
        # Abrir archivo procesamiento
        with open('./procesamiento.json', mode = 'r', encoding = 'utf8') as jsonfile:
            data = json.load(jsonfile)

        try:
            consolidation_file = data['ruta_consolidado_rechazos']

            return consolidation_file
        except KeyError:
            
            return None

    def _get_app_list(self):
        """Obtiene la lista de aprobadores de lecturas"""
        # Abrir archivo
        df = pd.read_excel(os.path.join(self.static, 'Aprobadores lecturas.xlsx'))

        # Ordenar
        df.sort_values('Orden', ascending = True, inplace = True)

        # Lista aprobadores
        app_list = df.iloc[:, 0].to_list()

        return app_list
    
    def _get_lectura_path(self):
        """Obtiene el ciclo de la lectura que se procesará"""
        # Definir carpeta donde se guardará la lectura
        lectura_folder = os.path.join(self.paths['carpeta_rechazos'], self.date.strftime('%Y'), str(self.date.month) + '. ' + self.date.strftime('%B').upper(), self.settings['lectura_manual']['nombre_carpeta'])

        # Validar si existe la carpeta
        if not os.path.exists(lectura_folder):
            os.makedirs(lectura_folder)
            lectura_cycle = '1'
        else:
            # Validar las lecturas guardadas en la carpeta
            files = os.listdir(lectura_folder)
            files = [f for f in files if re.search(self.date.strftime('%Y%m%d'), f) != None]

            if len(files) == 0:
                lectura_cycle = '1'
            else:
                cycles = [int(re.split('[_\.]', c)[2]) for c in files]
                lectura_cycle = str(max(cycles) + 1)

        lectura_file_name = self.settings['lectura_manual']['nombre'].replace('AAAAMMDD_#', self.date.strftime('%Y%m%d') + '_' + lectura_cycle)

        return os.path.join(lectura_folder, lectura_file_name)

    def config_driver(self, wait):
        """Configura el driver con el cual se realizará la lectura manual"""
        download_path = os.path.realpath(os.path.abspath('./temp'))
        settings = {

            'recentDestinations': [
                {
                    'id': 'Save as PDF',
                    'origin': 'local',
                    'account': '',
                }
                ],

            'selectedDestinationId': 'Save as PDF',
            'version': 2
        }

        options = Options()
        options.add_experimental_option('excludeSwitches', ['enable-logging', 'enable-automation']) # Quitar logs y que aparezca que está siendo usado por un software de pruebas
        options.add_experimental_option("detach", True)
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        options.add_argument("--start-maximized")
        options.add_argument("--disable-popup-blocking")
        options.add_argument('--kiosk-printing')

        options.add_experimental_option('prefs', {
            'download.default_directory': download_path,
            'savefile.default_directory': download_path,
            'credentials_enable_service': False,
            'printing.print_preview_sticky_settings.appState': json.dumps(settings)
        })

        # Iniciar driver
        driver = webdriver.Chrome(service = Service(), options = options)

        # Espera
        wait = WebDriverWait(driver, wait)

        return driver, wait

    def read_rej_sheet(self):
        """Carga la información que se encuentra en el consolidado del día"""
        # Definir ruta archivo
        self.rej_folder = os.path.join(self.paths['carpeta_rechazos'], self.date.strftime('%Y'), str(self.date.month) + '. ' + self.date.strftime('%B').upper(), self.settings['consolidado_rechazos']['nombre_carpeta'])
        self.consolidation_file = os.path.join(self.rej_folder, self.settings['consolidado_rechazos']['nombre'].replace('AAAAMMDD', self.date.strftime('%Y%m%d')))

        # Cargar hoja de rechazos en Dataframe
        df = pd.read_excel(self.consolidation_file, sheet_name = self.settings['consolidado_rechazos']['hoja_rechazos'])

        # Filtrar
        df = df[((df['Clasificación'].isnull()) | (df['Clasificación'] == 'EMBARGOS LECTURA TIEMPO REAL') | (df['Clasificación'] == 'TRN SUBSIDIO')) & (df['Estado rechazo'] != 'APLICADO')].reset_index(drop = True)
        self.df_soporte = df
        self.df_soporte.sort_values(by = ['Código DB o CR'], inplace = True)

        if df.shape[0] > 0:
            # Ajustar longitud nombre trx
            df['Descripción TRN'] = df['Descripción TRN'].str.slice(0, 29)

            # Ajustar fecha efectiva
            df['Fecha efectiva (AAAAMMDD)'] = df.apply(lambda row: int(self.date.strftime('%Y%m%d')) if datetime.strptime(str(row['Fecha efectiva']), '%Y%m%d') < datetime.now() + relativedelta(months = -1) or row['Código devolución'] == 'T' else row['Fecha efectiva'], axis = 1)

            # Ajustar monto créditos
            df['Monto'] = df.apply(lambda row: abs(row['Monto']), axis = 1)

            # Definir campos con texto fijo
            df['Observaciones'] = "Aplicacion rechazo depositos"

            # Definir columna vacía nombre trx
            df['Nombre de la transacción'] = pd.Series(dtype = str)

            # Cambiar cod trx generica ahorros debito
            idxs = df[(df['Código TRN rechazo'].astype(int) == 0) & (df['Código aplicación'] == 'S') & (df['Código DB o CR'] == 'D')].index
            df.loc[idxs, 'Código TRN rechazo'] = 1849
            df.loc[idxs, 'Nombre de la transacción'] = df.loc[idxs, 'Descripción TRN']

            # Cambiar cod trx generica ahorros crédito
            idxs = df[(df['Código TRN rechazo'].astype(int) == 0) & (df['Código aplicación'] == 'S') & (df['Código DB o CR'] == 'C')].index
            df.loc[idxs, 'Código TRN rechazo'] = 1587
            df.loc[idxs, 'Nombre de la transacción'] = df.loc[idxs, 'Descripción TRN']

            # Cambiar cod trx generica cte debito
            idxs = df[(df['Código TRN rechazo'].astype(int) == 0) & (df['Código aplicación'] == 'D') & (df['Código DB o CR'] == 'D')].index
            df.loc[idxs, 'Código TRN rechazo'] = 1761
            df.loc[idxs, 'Nombre de la transacción'] = df.loc[idxs, 'Descripción TRN']

            # Cambiar cod trx generica cte crédito
            idxs = df[(df['Código TRN rechazo'].astype(int) == 0) & (df['Código aplicación'] == 'D') & (df['Código DB o CR'] == 'C')].index
            df.loc[idxs, 'Código TRN rechazo'] = 3681
            df.loc[idxs, 'Nombre de la transacción'] = df.loc[idxs, 'Descripción TRN']

            # Cambiar cod trx generica embargos cte
            idxs = df[(df['Clasificación'] == 'EMBARGOS LECTURA TIEMPO REAL') & (df['Código aplicación'] == 'D')].index
            df.loc[idxs, 'Código TRN rechazo'] = 9900
            df.loc[idxs, 'Nombre de la transacción'] = df.loc[idxs, 'Descripción TRN']

            # Cambiar cod trx generica embargos ahorros
            idxs = df[(df['Clasificación'] == 'EMBARGOS LECTURA TIEMPO REAL') & (df['Código aplicación'] == 'S')].index
            df.loc[idxs, 'Código TRN rechazo'] = 9999
            df.loc[idxs, 'Nombre de la transacción'] = df.loc[idxs, 'Descripción TRN']

            # Dejar solamente columnas que se necesitan para la lectura
            lect_cols = self.settings['lectura_manual']['columnas']
            cols = [c for c in df.columns if c in lect_cols.values()]
            df= df[cols]

            # Cambiar nombres a las columnas
            new_col_names = {k:v for v, k in lect_cols.items() if k in cols}
            df.rename(columns = new_col_names, inplace = True)

            # Ordernar por naturaleza crédito
            df.sort_values(by = ['Número de cuenta' ,'Naturaleza\nC: Crédito\nD: Débito'], inplace = True)
            
            self.df_lectura = df.reset_index(drop = True)
            self.df_lectura.to_excel('./out/lectura.xlsx', index = False)
        else:
            msg = 'No se encontraron registros para realizar lectura manual'
            self.show_messages(msg, 'warning', True)
        
        return None

    def gen_lectura_file(self):
        """Guarda archivo lectura manual en la carpeta del ciclo realizado"""
        self.logger.info('Generando planilla lectura')

        # Copiar plantilla de lectura en ruta
        shutil.copy(os.path.join(self.static, self.settings['lectura_manual']['plantilla']), self.lectura_path)

        # Abrir archivo lectura y definir hoja
        wb = xl.load_workbook(self.lectura_path)
        ws = wb[self.settings['lectura_manual']['hoja_lectura']]

        # Definir formatos para celdas
        cells_format = {}

        for i, cell in enumerate(ws[2], 1):
            if cell.has_style:
                col = ws.cell(1, i).value
                cells_format[col] = {
                    'font': copy(cell.font),
                    'border': copy(cell.border),
                    'fill': copy(cell.fill),
                    'number_format': copy(cell.number_format),
                    'protection': copy(cell.protection),
                    'alignment': copy(cell.alignment)
                }

        # Llenar archivo
        self.fill_excel_sheet(ws, self.df_lectura, cells_format = cells_format)

        # Guardar soporte
        wb.create_sheet('SOPORTE')
        sheet = wb['SOPORTE']

        for j, column in enumerate(self.df_soporte.columns, 1):
            cell = sheet.cell(row = 1, column = j)
            cell.value = column
        
        self.fill_excel_sheet(sheet, self.df_soporte)

        # Guardar
        wb.save(self.lectura_path)

        self.logger.info('Lectura guardada en ' + self.lectura_path)

        return None
    
    def _open_lect_url(self, driver, wait):
        """Abre el sitio de lecturas realizando la autenticación"""
        # Intentar abrir sitio de lecturas hasta 3 veces
        max_tries = 3
        n = 1

        while n <= max_tries:
            self.logger.debug('Abriendo sitio lecturas intento ' + str(n) + ' de ' + str(max_tries))
            try:
                # Abrir sitio
                driver.get(self.config['url_lecturas'])

                # Digitar usuario y contraseña
                var = True

                while var:
                    try:
                        wait.until(EC.presence_of_element_located((By.XPATH, self.xpaths['usuario_red']))).send_keys(self.red_user + Keys.ENTER)
                        sleep(1)
                        break
                    except ElementNotInteractableException:
                        pass
                    except StaleElementReferenceException:
                        pass

                while var:
                    try:
                        wait.until(EC.presence_of_element_located((By.XPATH, self.xpaths['contraseña_red']))).send_keys(self.red_pwd)
                        sleep(1)
                        wait.until(EC.element_to_be_clickable((By.XPATH, '//input[@type="submit"]'))).click()
                        sleep(1)
                        break
                    except ElementNotInteractableException:
                        pass
                    except StaleElementReferenceException:
                        pass

                # Esperar hasta tres veces el tiempo parametrizado en wait
                for i in range(1, 4):
                    try:
                        wait.until(EC.element_to_be_clickable((By.XPATH, self.xpaths['entrar_btn'])))
                        sleep(1)
                        self.driver.find_element(By.XPATH, self.xpaths['entrar_btn']).click()
                        break
                    except TimeoutException:
                        if i == 3:
                            raise Exception
                
                self.driver, self.wait = driver, wait
                break
            except:
                driver.close()
                if n == max_tries:
                    raise Exception
                else:
                    n += 1
                    driver, wait = self.config_driver(20)
                
        return None
    
    def rad_lectura(self):
        """Radica el archivo de lectura en el sitio"""
        self.logger.info('Radicando lectura')

        # Instanciar chrome
        self.driver, self.wait = self.config_driver(60)

        # Abrir sitio lecturas y entrar
        self._open_lect_url(self.driver, self.wait)
        
        # Radicar movimiento manual
        self.wait.until(EC.element_to_be_clickable((By.XPATH, self.xpaths['radicar_mvto_btn']))).click()
        sleep(1)

        # Adjuntar archivo
        input = self.wait.until(EC.presence_of_element_located((By.XPATH, self.xpaths['adjuntar_archivo'])))
        input.send_keys(os.path.realpath(self.lectura_path))
        # input.send_keys(os.path.realpath(r"C:\Users\dcastano\Downloads\Radicacion de movimientos.xlsx"))
        sleep(1)

        # Validar carga de archivo exitosa
        var = True

        while var:
            try:
                self.driver.find_element(By.XPATH, self.xpaths['carga_exitosa'])
                sleep(1)
                break
            except NoSuchElementException:
                pass
        
        # Poner comentarios y continuar
        self.driver.find_element(By.XPATH, self.xpaths['comentarios']).send_keys('Aplicación rechazos depósitos')
        self.driver.find_element(By.XPATH, self.xpaths['continuar_btn']).click()

        # Obtener número radicado
        self.wait.until(EC.presence_of_element_located((By.XPATH, self.xpaths['numero_radicado'])))
        sleep(2)
        self.rad_number = self.driver.find_element(By.XPATH, self.xpaths['numero_radicado']).text
        self.logger.info('Radicado lectura: ' + self.rad_number)

        # Obtener total de registros y monto total
        tot_rec = int(self.driver.find_element(By.XPATH, self.xpaths['total_registros']).text)
        tot_value = float(re.sub(r'[\$\,]', '', self.driver.find_element(By.XPATH, self.xpaths['total_monto']).text))
        self.logger.info('Total registros lectura: ' + str(tot_rec))
        self.logger.info('Total monto lectura: ' + self.driver.find_element(By.XPATH, self.xpaths['total_monto']).text)

        # Realizar validación de registros y monto
        if (tot_rec != self.df_lectura.shape[0]) or (tot_value != round(self.df_lectura['Valor de la transacción'].astype(float).sum(), 2)):
             msg = 'El total de registros o el total del monto no coincide con el de la lectura, Total registros: ' + str(self.df_lectura.shape[0]) + ' Total valor: ' + str(self.df_lectura['Valor de la transacción'].astype(float).sum())
             raise Exception(msg)
        
        # Seleccionar aprobador 
        app_list = []

        while len(app_list) == 0:
            self.wait.until(EC.element_to_be_clickable((By.XPATH, self.xpaths['seleccionar_aprobador'])))
            sleep(2)
            self.driver.find_element(By.XPATH, self.xpaths['seleccionar_aprobador']).click()
            sleep(2)
            app_list = self.driver.find_elements(By.XPATH, self.xpaths['lista_aprobadores'])

        app_selected = None

        # Iterar la lista hasta encontrar aprobador
        for approver in self.app_list:
            try:
                app_selected = [a for a in app_list if a.text == approver][0]
                app_selected.click()
                sleep(1)
                self.logger.info('Aprobador seleccionado: ' + approver)
                break
            except IndexError:
                pass
        
        # Validar si se encontró aprobador
        if app_selected == None:
            msg = 'No se encontró aprobador de lectura'
            raise Exception(msg)
        else:
            # Finalizar radicación lectura
            self.wait.until(EC.element_to_be_clickable((By.XPATH, self.xpaths['finalizar_btn']))).click()

            # Guardar datos de lectura
            self.update_json(**{'radicado_lectura': self.rad_number, 'ruta_consolidado_rechazos': self.consolidation_file})
        
        return None
    
    def wait_mail_confirmation(self):
        """Espera la confirmación de la gestión de la lectura"""
        self.logger.info('Esperando mail de respuesta de gestión de lectura radicado número: ' + self.rad_number)

        # Crear instancia de outlook y extraer mensajes en la bandeja de entrada
        outlook = win32com.Dispatch("Outlook.Application").GetNamespace("MAPI")
        
        # Inicializar variable de mail recibido
        mail_recieved = False

        while not mail_recieved:
            inbox = outlook.GetDefaultFolder(6) 
            messages = inbox.Items

             # Recorrer cada mensaje de la bandeja y dejar solo los mensajes de hoy que el asunto haga match con el especificado
            for message in messages:
                if re.search(self.settings['lectura_manual']['asunto_mail'].replace('#ID', self.rad_number), message.Subject, re.IGNORECASE) != None and message.Senton.date() == datetime.now().date():
                    mail_recieved = True
                    self.logger.info(f'Mail de solicitud {self.rad_number} de lectura gestionada recibido')
                    break

    def search_rad(self):
        """Busca y descarga solicitud"""
        self.logger.info('Descargando lectura radicado: ' + self.rad_number)

        # Instanciar chrome si no se encuentra abierto
        if self.driver == None:
            self.driver, wait = self.config_driver(20)
        else:
            wait = WebDriverWait(self.driver, 20)

        # Limpiar carpeta temporal
        for file in os.listdir('./temp'):
            os.remove(os.path.join('./temp', file))

        # Crear carpeta de respuesta lectura
        lectura_reply_folder = os.path.join(self.paths['carpeta_rechazos'], self.date.strftime('%Y'), str(self.date.month) + '. ' + self.date.strftime('%B').upper(), self.settings['lectura_manual']['carpeta_respuesta'])

        if not os.path.exists(lectura_reply_folder):
            os.makedirs(lectura_reply_folder)

        # Abrir sitio lecturas y entrar
        self._open_lect_url(self.driver, wait)

        # Consultar solicitud 
        wait.until(EC.element_to_be_clickable((By.XPATH, self.xpaths['consultar_solicitudes']))).click()
        sleep(2)

        # Digitar radicado
        self.driver.find_element(By.XPATH, self.xpaths['radicado_solicitud_textbox']).send_keys(self.rad_number)
        sleep(1)

        # Darc clic en buscar
        wait.until(EC.element_to_be_clickable((By.XPATH, self.xpaths['buscar_solicitud']))).click()
        sleep(2)

        # Validar el estado de la solicitud
        try:
            rad_state = wait.until(EC.presence_of_element_located((By.XPATH, self.xpaths['estado_solicitud'])))
        except TimeoutException:
            msg = 'No se encontró lectura con radicado número: ' + self.rad_number
            raise Exception(msg)

        sleep(2)
        self.rad_state = rad_state.text
        self.logger.info('Estado de la solicitud de lectura ------> ' + self.rad_state)

        # Mostrar mensaje si el estado es diferente al aplicado
        if self.rad_state != 'Aplicado':
            msg = 'Lectura no procesada: Por favor corregir inconsistencias'
            raise Exception(msg)

        # Descargar solicitud
        wait.until(EC.element_to_be_clickable((By.XPATH, self.xpaths['descargar_solicitud']))).click()
        sleep(10)

        # Esperar a que termine de descargar
        var = True

        while var:
            while len(os.listdir(os.path.abspath('./temp'))) == 0:
                pass

            for file in os.listdir(os.path.abspath('./temp')):
                if re.match(self.rad_number + '_Portal transaccional', file) and os.path.splitext(file)[1] == '.xlsx':
                    self.lectura_reply_path = os.path.join(lectura_reply_folder, file)
                    shutil.move(os.path.join(os.path.abspath('./temp'), file), self.lectura_reply_path)
                    var = False
        
        # Cerrar driver
        self.driver.close()

        return None
    
    def update_consolidation_file(self):
        """Actualizar archivo de consolidación de acuerdo a la respuesta de la lectura"""
        # Abrir consolidado de rechazos
        df = pd.read_excel(self.consolidation_file, sheet_name = self.settings['consolidado_rechazos']['hoja_rechazos'])

        # Abrir respuesta lectura
        df_lectura_reply = pd.read_excel(self.lectura_reply_path)

        # Hacer merge entre consolidado
        df = df.merge(df_lectura_reply, how = 'left', left_on = 'llave', right_on = 'Detalle o información adicional (F0210)', validate = 'm:1')

        # Validar respuestas exitosas EMBARGOS
        idxs = df.loc[(df['Respuesta servicio'] == self.config['respuestas_servicio_lectura']['exitoso']) & (df['Clasificación'] == 'EMBARGOS LECTURA TIEMPO REAL')].index
        df.loc[idxs, 'Clasificación'] = 'EMBARGOS LECTURA TIEMPO REAL'
        df.loc[idxs, 'Estado rechazo'] = 'APLICADO'

        # Validar respuestas exitosas SUBSIDIO
        idxs = df.loc[(df['Respuesta servicio'] == self.config['respuestas_servicio_lectura']['exitoso']) & (df['Clasificación'] == 'TRN SUBSIDIO')].index
        df.loc[idxs, 'Clasificación'] = 'TRN SUBSIDIO LECTURA TIEMPO REAL'
        df.loc[idxs, 'Estado rechazo'] = 'APLICADO'

        # Validar respuestas exitosas
        idxs = df.loc[(df['Respuesta servicio'] == self.config['respuestas_servicio_lectura']['exitoso']) & (df['Clasificación'].isnull())].index
        df.loc[idxs, 'Clasificación'] = 'LECTURA TIEMPO REAL'
        df.loc[idxs, 'Estado rechazo'] = 'APLICADO'

        # Validar respuestas no exitosas
        idxs = df.loc[df['Respuesta servicio'] == self.config['respuestas_servicio_lectura']['error']].index
        df.loc[idxs, 'Estado rechazo'] = 'PENDIENTE'

        # Actualizar archivo consolidado
        self.update_excel_file(self.consolidation_file, self.settings['consolidado_rechazos']['hoja_rechazos'], df)

        # Actualizar json
        self.update_json(**{'radicado_lectura': None})

        return None
    
    def run(self):
        """Orquesta la ejecución de la clase"""
        if self.rad_number == None:
            self.read_rej_sheet()
            self.gen_lectura_file()
            self.rad_lectura()
            
        self.wait_mail_confirmation()
        self.search_rad()
        self.update_consolidation_file()

        return None