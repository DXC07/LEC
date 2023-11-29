from logger import Logging
import json
from datetime import datetime
import warnings
import pyodbc
import re
import traceback
import os
import locale
import tkinter as tk
from tkinter import messagebox
import pandas as pd
from pandas.io.sql import DatabaseError
import openpyxl as xl

# Definir región fecha
locale.setlocale(locale.LC_ALL, 'es_CO')

# Cargar archivo de configuración
with open('./static/config.json', encoding='utf-8') as f:
    config = json.load(f)

# Deshabilitar advertencias
warnings.filterwarnings('ignore')


class Functions:

    def __init__(self):
        self.today = datetime.now() # Día de hoy
        self.credentials = None # Usuarios y contraseñas
        self.static = os.path.abspath('./static') # Carpeta Static

        # Cargar archivo de configuración
        with open(os.path.join(self.static, 'config.json'), encoding='utf-8') as f:
            self.config = json.load(f)
        
        # Cargar archivo de ajustes
        with open(os.path.join(self.static, 'settings.json'), encoding='utf-8') as f:
            self.settings = json.load(f)

        self.paths = self.config['rutas'] # Rutas
        self.sql_path = os.path.join(self.static, 'sql')
        self.logger = Logging().get_logger(self.__module__, self.paths['log_ejecuciones']) # Logger
        self.rej_folder = None # Carpeta donde se guardará el consolidado
        self.consolidation_file = None # Ruta del archivo consolidado con el query de rechazos
    
    def show_messages(self, msg, type, log = False):
        """Muestra mensajes de información, error y advertencias"""
        app = tk.Tk()
        app.withdraw()

        if type == 'info':
            if log:
                self.logger.info(msg)
            messagebox.showinfo('Info', msg)
        elif type == 'error':
            if log:
                self.logger.error(msg)
            messagebox.showerror('Error', msg)
        elif type == 'warning':
            if log:
                self.logger.warning(msg)
            messagebox.showwarning('Advertencia', msg)

        return None

    @staticmethod
    def save_execution_time(func):
        """Decorador para guardar el tiempo de ejecución de una función"""
        def wrapper(self, *args, **kwargs):
            initial_time = datetime.now()
            func(self, *args, **kwargs)
            final_time = datetime.now()
            time_elapsed = final_time - initial_time
            
            # Guardar
            try:
                times_path = os.path.join(self.paths['log_ejecuciones'], 'Tiempo ejecuciones.xlsx')
                df = pd.read_excel(times_path, dtype = str)
                df['Fecha'] = pd.to_datetime(df['Fecha'], format = '%Y-%m-%d')
            except:
                df = pd.DataFrame(columns = ['Herramienta', 'Fecha', 'Usuario', 'Módulo', 'Tiempo ejecución' ])
            
            df.loc[len(df.index)] = [self.settings['nombre_euc'], datetime.now(), os.getlogin(), func.__name__, str(time_elapsed)]
            df.to_excel(times_path, index = False)

            self.show_messages('Fin ' + func.__name__, 'info', log = True)
            self.logger.info(f'Tiempo ejecucion {func.__name__}: {time_elapsed}')

        return wrapper

    @staticmethod
    def log_errors(func):
        """Decorador para capturar y lanzar errores de ejecución"""
        def wrapper(self, *args, **kwargs):
            try:
                return func(self, *args, **kwargs)

            except Exception as e:
                msg = str(e) + ' - ' + str(traceback.format_exc())
                self.logger.error(msg)
                self.show_messages(msg, 'error')
                exit()
        
        return wrapper

    def fill_excel_sheet(self, sheet, df, row = 2, cells_format = None):
        """Llena la información de una hoja de excel"""
        # Encabezados
        headers = [cell.value for cell in sheet[1]]

        # Llenar hoja
        for col, header in enumerate(headers, 1):
            for i in df.index:
                cell = sheet.cell(row = i + row, column = col)

                # Digitar valor
                if header in df.columns:
                    cell.value = df[header][i]

                # Aplicar formato
                if cells_format != None:
                    try:
                        _format = cells_format[header]
                    except KeyError:
                        pass
                    else:
                        for k, v in _format.items():
                            setattr(cell, k, v)
                    
        return None
    
    def update_excel_file(self, consolidation_file, sheet_name, df, clean_sheet = True, row = 2, cells_format = None):
        """Actualiza archivo excel"""
        # Abrir archivo
        wb = xl.load_workbook(consolidation_file)

        # Definir hoja
        ws = wb[sheet_name]

        # Limpiar hoja
        if clean_sheet:
            for i in range(row, ws.max_row + 1):
                for j in range(1, ws.max_column + 1):
                    ws.cell(row = i, column = j).value = None

        # Llenar hojas
        self.fill_excel_sheet(ws, df, row, cells_format)

        # Guardar
        wb.save(consolidation_file)

        return None

    def connect_odbc(self, application, **kwargs):
        """Retorna conexión a un driver ODBC"""

        if application == 'ACCESS':
            db_path = kwargs['db_path']
            conn = pyodbc.connect('Driver={%s};DBQ=%s;' % (self.settings['drivers_odbc']['access'], db_path), autocommit = True)
            cursor = conn.cursor()

            return conn, cursor
        elif application == 'AS400':
            maq = kwargs['maq']
            user = kwargs['user']
            pwd = kwargs['pwd']

            if maq == 'NACIONAL':
                # conn = pyodbc.connect('driver={%s};system=10.9.2.201;UID={%s};PWD={%s};Trusted_Connection=no;autocommit=True' % (self.settings['drivers_odbc']['AS400'], str(user), str(pwd)), autocommit = True)
                conn = pyodbc.connect(f'DSN=NACIONAL;UID={user};PWD={pwd};Trusted_Connection=no;autocommit=True', autocommit = True)
            elif maq == 'MEDELLIN':
                conn = pyodbc.connect('driver={%s};system=MEDELLINET01;UID={%s};PWD={%s};Trusted_Connection=no' % (self.settings['drivers_odbc']['AS400'], str(user), str(pwd)))
            else:
                raise Exception('La máquina especificada no es válida')
            
            cursor = conn.cursor()

            return conn, cursor
        else:
            raise Exception('Aplicación no válida')
    
    def query_AS400(self, sql, conn, raise_error = False):
        """Ejecuta un query dada una sentencia SQL y una conexión y retorna un dataframe con los resultados"""
        self.logger.debug('Query a ejecutar:\n\t\t' + sql.replace('\n', '\n\t\t') + '\n')

        # Intentar ejecutar el query y el estado de la conexión
        try:
            df = pd.read_sql_query(sql, conn)
            e = 'Conexión OK'
            return df, e
        except DatabaseError as e:
            if raise_error:
                raise Exception
            else:
                return pd.DataFrame(), e
            
    def val_permission(self):
        """Valida acceso a los objetos y librerías especificados"""

        self.logger.info('Validando accesos a librerías')
        errors = []
        msg = ''

        # Cargar objetos
        objs = self.settings['objetos_AS400']

        # Recorrer cada objeto y validar si es exitosa la conexión
        for obj, maq in objs.items():
            self.logger.debug(f'Validando objeto {obj} en máquina {maq}')

            sql = f"""SELECT * FROM {obj} LIMIT 1"""

            if maq == 'NACIONAL':
                user = self.credentials['login_NACIONAL_user']
                pwd = self.credentials['login_NACIONAL_pwd']
            else:
                user = self.credentials['login_MEDELLIN_user']
                pwd = self.credentials['login_MEDELLIN_pwd']

            conn = self.connect_odbc('AS400', user = user, pwd = pwd, maq = maq)[0]
            e = str(self.query_AS400(sql, conn)[1])
            
            if e != 'Conexión OK':
                errors.append((obj, e))
                self.logger.error(f'Validación de objeto {obj} AS400 NO exitosa')
            else:
                self.logger.info(f'Validación de objeto {obj} AS400 exitosa')
        
        # Renombrar errores conocidos
        if len(errors) > 0:
            for obj, e in errors:
                if re.search('Not authorized', e) != None:
                    error = 'No tiene autorización a este objeto.'
                elif re.search('FILE not found', e) != None:
                    error = 'cuentra este objeto.'
                else:
                    error = e

                msg += obj + ': ' + error + '\n'
            
            self.show_messages(msg, 'error')

            raise Exception('No fue posible hacer conexión con las librerías')

        return None
        
    def update_json(self, **kwargs):
        """Actualiza archivo de procesamiento"""
        file = os.path.abspath('./procesamiento.json')

        # Leer archivo
        with open(file, "r", encoding = 'utf-8') as jsonfile:
            data = json.load(jsonfile)
        
        # Actualizar diccionario
        for k, v in kwargs.items():
            data[k] = v

        # Sobrescribir archivo
        with open(file, "w", encoding = 'utf-8') as jsonfile:
            json.dump(data, jsonfile, indent = 4, separators = (',', ': '), ensure_ascii = False)
        
        # Cerrar archivo
        jsonfile.close()

        return None

    def send_mail(self, outlook, to, subject, display = True, send = False, **kwargs):
        """Envía mail con el emisor, destinatarios y adjuntos especificados"""
        mail = outlook.CreateItem(0)
        mail.To = to
        mail.Subject = subject

        # Enviar con otro correo email
        try:
            mail.SentOnBehalfOfName = kwargs['acc']
        except KeyError:
            pass

        # Ingresar cuerpo con texto sin formato
        try:
            mail.Body = kwargs['body']
        except KeyError:
            pass
        
        # Ingresar cuerpo con HTML
        try:
            mail.HTMLBody = kwargs['hbody']
        except KeyError:
            pass
        
        # Adjuntar archivos
        try:
            for attachment in kwargs['attachments']:
                mail.Attachments.Add(attachment)
        except KeyError:
            pass
        
        # Mostrar el mail
        if display:
            mail.Display()
        
        # Enviar mail automáticamente
        if send:
            mail.Send()

        return None