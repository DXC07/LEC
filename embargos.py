from functions import Functions
import os
import pandas as pd
import io
from datetime import datetime
import shutil
import json
import win32com.client as win32com
import re
import numpy as np
import openpyxl as xl
from openpyxl.utils import get_column_letter
import tkinter as tk
class Embargos(Functions):

    def __init__(self, credentials, date, window):
        super().__init__()
        self.file_sent = self._get_file_flag() # Marcar para identificar si ya se radicó a embargos
        self.nac_user = credentials['login_NACIONAL_user'] # Usuario nacional
        self.nac_pwd = credentials['login_NACIONAL_pwd'] # Contraseña nacional
        self.window = window # Ventana tkinter
        self.conn = self.connect_odbc('AS400', maq = 'NACIONAL', user = self.nac_user, pwd = self.nac_pwd)[0] # Conexión ODBC Nacional
        self.date = date # Fecha del último día de los rechazos gestionados
        self.rej_folder = os.path.join(self.paths['carpeta_rechazos'], self.date.strftime('%Y'), str(self.date.month) + '. ' + self.date.strftime('%B').upper(), self.settings['consolidado_rechazos']['nombre_carpeta'])
        self.consolidation_file = os.path.join(self.rej_folder, self.settings['consolidado_rechazos']['nombre'].replace('AAAAMMDD', self.date.strftime('%Y%m%d')))
        self.df_emb_reply = None # Respuesta de embargos
        self.send = True if self.file_sent else False # Bandera para identificar si el usuario quiere radicar a embargos
    
    def _get_file_flag(self):
        """Obtiene la marca de si ya se envió el archivo embargos"""
        # Abrir archivo procesamiento
        with open('./procesamiento.json', mode = 'r', encoding = 'utf8') as jsonfile:
            data = json.load(jsonfile)

        try:
            file_sent = data['archivo_embargos_radicado']

            return file_sent
        except KeyError:
            
            return None
        
    def _get_trn_hour(self, df_trns):
        """Obtiene la hora de las transacciones enviadas"""
        # Hacer reset del índice
        df_trns = df_trns.reset_index(drop = False).rename(columns = {'index': 'indice'})

        # Cambiar formato a fecha
        df_trns['Fecha efectiva'] = pd.to_datetime(df_trns['Fecha efectiva'].astype(str), format = '%Y%m%d')

        # Cargar consulta sql itc
        with io.open(os.path.join(self.sql_path, 'itc.sql'), mode = 'r', encoding = 'utf8') as f:
            sql_itc = f.read()

        # Recorrer dataframe
        for i in df_trns.index:
            self.logger.info('Buscando trx embargo ' + str(i + 1) + ' de ' + str(df_trns.shape[0]))
            # Extraer valores
            nit = "'" + ('0' * 15 + str(int(df_trns['Número identificación'][i])))[-15:] + "'"
            account = int(df_trns['Número cuenta'][i])
            value = abs(df_trns['Monto'][i])
            date = df_trns['Fecha efectiva'][i]

            # Reemplazar valores en consulta
            sql = sql_itc.replace('nitcliente', str(nit))
            sql = sql.replace('nrocuenta', str(account))
            sql = sql.replace('valortrn', str(value))
            sql = sql.replace('AAAA', date.strftime('%Y'))
            sql = sql.replace('MM', date.strftime('%m'))
            sql = sql.replace('DD', date.strftime('%d'))

            # Ejecutar query
            df = self.query_AS400(sql, self.conn, True)[0]

            # Obtener hora transacción
            if df.shape[0] > 0:
                trn_hour = datetime.strptime(('0' * 8 + str(int(df['HORATRN'][0])))[-8:], '%H%M%S%f').strftime('%H:%M:%S')
                df_trns.loc[i, 'Hora TRN'] = trn_hour
        
        # Definir como indice la columna indice
        df_trns.set_index('indice', inplace = True)

        return df_trns

    def val_send_file(self, tot_trn, tot_trn_hour):
        """Ventana para preguntar al usuario si quiere radicar a embargos"""
        def yes_action():
            self.send = True
            root.quit()
            
            return None

        def no_action():
            self.send = False
            root.quit()
            self.logger.warning('No se realizó radicación en el flujo de embargos')

            return None

        # Crear ventana
        root = tk.Toplevel(self.window)
        root.protocol("WM_DELETE_WINDOW", no_action)
        root.configure(bg="white")
        p1 = tk.PhotoImage(file = './static/Imágenes/logo.png')
        root.iconphoto(False, p1)
        root.title("")

        msg = 'Se encontró hora de ' + str(tot_trn_hour) + ' de ' + str(tot_trn) + ' trx de embargos'
        self.logger.info(msg)

        label = tk.Label(root, text = msg + '\n¿Desea radicar en el flujo?', bg = 'white', font = ('CIBFont Sans', 12))
        frame = tk.Frame(root, bg = 'white')
        yes_button = tk.Button(frame, text = "SI", command = lambda: yes_action(), width = 10, bg = '#2C2A29', fg = 'white', font = ('CIBFont Sans', 12))
        yes_button.pack(padx = 10, pady = 10, side = 'left')

        no_button = tk.Button(frame, text = "NO", command = no_action, width = 10, bg = '#2C2A29', fg = 'white', font = ('CIBFont Sans', 12))
        no_button.pack(padx = 10, pady = 10, side = 'left')

        label.pack(padx=20, pady=20, fill="both", expand=True)
        frame.pack()

        root.mainloop()
        root.destroy()

        return None

    def update_consolidation_file(self, emb_reply = False):
        """Actualiza el archivo de consolidado para enviar el archivo a embargos y luego de la respuesta"""
        # Cargar hoja de rechazos en Dataframe
        df = pd.read_excel(self.consolidation_file, sheet_name = self.settings['consolidado_rechazos']['hoja_rechazos'])
        
        if emb_reply == False:
            # Definir áreas
            areas = ['PAGOS', 'MEDIOS DE PAGO', 'RECAUDOS', 'CANJE', 'FINANCIACION']

            for area in areas:
                # Actualizar campo clasificación y estado rechazo
                idxs = df[(df['Clasificación'] == 'EMBARGOS') & (df[area].notnull())].index
                df.loc[idxs, 'Clasificación'] = area
                df.loc[idxs, 'Estado rechazo'] = 'CONTABILIZAR'

            # Buscar hora transacción
            df_trns = df[df['Clasificación'] == 'EMBARGOS'][['Número cuenta', 'Número identificación', 'Monto', 'Hora TRN', 'Fecha efectiva']]
            df_trns = self._get_trn_hour(df_trns)

            # Actualizar en dataframe de consolidado
            idxs = df_trns.index
            df.loc[idxs, 'Hora TRN'] = df_trns.loc[idxs, 'Hora TRN']

            self.df_rej = df
        else:
            # Cuentas que autorizaron
            # Hacer merge con archivo respuesta embargos
            df = pd.merge(df, self.df_emb_reply, how = 'left', on = 'llave', validate= 'm:1')

            # Actualizar clasificación
            idxs = df[df['¿Se autoriza?'].str.strip().str.upper().isin(['SI', 'SÍ'])].index
            df.loc[idxs, 'Clasificación'] = 'EMBARGOS LECTURA TIEMPO REAL'
            df.loc[idxs, 'Estado rechazo'] = 'PENDIENTE'

            # Clasificar rechazos embargos no autorizados
            idxs = df[(~df['¿Se autoriza?'].str.strip().str.upper().isin(['SI', 'SÍ'])) & (df['Clasificación'] == 'EMBARGOS')].index
            df.loc[idxs, 'Clasificación'] = "CXC"
            df.loc[idxs, 'Estado rechazo'] = "CONTABILIZAR"

            df.to_excel('./out/respuesta_emb.xlsx', index = False)

        # Actualizar archivo consolidado
        self.update_excel_file(self.consolidation_file, self.settings['consolidado_rechazos']['hoja_rechazos'], df, False)

        return None

    def gen_emb_file(self):
        """Genera el archivo que se enviará a embargos"""
        # Crear carpeta de embargos
        embargos_folder = os.path.join(self.paths['carpeta_rechazos'], self.date.strftime('%Y'), str(self.date.month) + '. ' + self.date.strftime('%B').upper(), self.settings['embargos']['nombre_carpeta'])

        if not os.path.exists(embargos_folder):
            os.makedirs(embargos_folder)
        
        embargos_file_name = self.settings['embargos']['nombre'].replace('AAAAMMDD', self.date.strftime('%Y%m%d'))
        
        # Ruta archivo
        self.embargos_file_path = os.path.join(embargos_folder, embargos_file_name)

        # Filtrar dataframe de consolidado por los rechazos de embargos
        df_embargos = self.df_rej[self.df_rej['Clasificación'] == 'EMBARGOS']

        if df_embargos.shape[0] > 0:
            # Contar a cuantas trns se les encontró la hora
            tot_trn = df_embargos.shape[0]
            tot_trn_hour = df_embargos[df_embargos['Hora TRN'].notnull()].shape[0]

            # Dejar solo las columnas que se necesitan y crear las que se deben dejar vacías
            embargos_cols = self.settings['embargos']['columnas']

            df_embargos = df_embargos[[c for c in df_embargos.columns if c in embargos_cols.values()]]

            for new_col, col in embargos_cols.items():
                if col in df_embargos.columns:
                    df_embargos.rename(columns = {col: new_col}, inplace = True)
                else:
                    df_embargos[col] = pd.Series(dtype = str)
            
            # Ordenar por nit
            df_embargos.sort_values('NIT cliente', inplace = True)

            # Exportar a excel
            df_embargos.to_excel(self.embargos_file_path, index = False)

            # Obtener la máxima longitud de cada columna del archivo
            widths = {}

            for col in df_embargos.columns:
                max_w_h = len(col)
                max_w_c = df_embargos[col].astype(str).str.len().max()
                widths[col] = max(max_w_h, max_w_c)

            # Actualizar ancho columnas
            wb = xl.load_workbook(self.embargos_file_path)
            ws = wb['Sheet1']

            for j, col in enumerate(df_embargos.columns, 1):
                ws.column_dimensions[get_column_letter(j)].width = widths[col] + 1

            # Guardar
            wb.save(self.embargos_file_path)

            # Validar si se quiere enviar
            self.val_send_file(tot_trn, tot_trn_hour)

            if self.send == False:
                return None

            self.logger.info('Radicando en flujo de embargos')

            # Copiar a carpeta de OneDrive para que Power Automate lo cargue en la lista de Sharepoint
            shutil.copy(self.embargos_file_path, os.path.join(self.paths['onedrive'], 'Embargos', embargos_file_name))

            # Actualizar archivo procesamiento
            self.update_json(archivo_embargos_radicado = True)
            self.update_json(llaves_embargos = df_embargos['llave'].to_list())
        else:
            msg = 'No se encontraron rechazos para enviar a embargos'
            self.show_messages(msg, 'warning', True)

        return None

    def wait_emb_file(self):
        """Copia en la carpeta de Onedrive"""
        self.logger.info('Esperando respuesta embargos')
        
        # Obtener llaves embargos
        with open('./procesamiento.json', mode = 'r', encoding = 'utf8') as jsonfile:
            data = json.load(jsonfile)

        emb_keys = data['llaves_embargos']
        
        # Asunto correo respuesta
        subject = 'Solicitud autorización rechazos depósitos'

        # Crear instancia de outlook y extraer mensajes en la bandeja de entrada
        outlook = win32com.Dispatch("Outlook.Application").GetNamespace("MAPI")
        inbox = outlook.GetDefaultFolder(6) 
        messages = inbox.Items

        # Inicializar variable
        found = False

        # Recorrer cada mensaje de la bandeja y dejar solo los mensajes de hoy que no estén leídos y que el asunto haga match con el especificado
        while found == False:
            for message in messages:
                # Dejar de buscar si ya se encontró el mensaje
                if found:
                    break

                if re.search(subject, message.Subject, re.IGNORECASE) != None and message.Senton.date() == datetime.now().date():
                    attachments = message.Attachments
                    for attachment in attachments:
                        if os.path.splitext(str(attachment))[1] == '.xlsx':
                            attachment_path = os.path.join(os.path.abspath('./temp'), str(attachment))
                            attachment.SaveAsFile(attachment_path)

                            # Abrir archivo para validar que si se haya dado respuesta
                            try:
                                df = pd.read_excel(attachment_path)

                                if df['¿Se autoriza?'].isnull().sum() < df.shape[0]:
                                    # Validar que el archivo contenga las mismas llaves enviadas
                                    keys_rep_file = df['llave'].to_list()

                                    if sorted(emb_keys) == sorted(keys_rep_file):
                                        self.df_emb_reply = df
                                        found = True
                                        break
                                    else:
                                        raise Exception
                                else:
                                    raise Exception
                            except:
                                os.remove(attachment_path)
        
        return None
    
    def run(self):
        """Orquesta la ejecución de la clase"""
        # Validar si ya se radicó el archivo a embargos
        if self.file_sent != True:
            self.update_consolidation_file()
            self.gen_emb_file()
        
        if self.send:
            self.wait_emb_file()
            self.update_consolidation_file(True)
        
        return None