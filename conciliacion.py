from functions import Functions
import io
import os
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill

class Conciliacion(Functions):
    def __init__(self, credentials, conc_date):
        super().__init__()
        self.nac_user = credentials['login_NACIONAL_user'] # Usuario nacional
        self.nac_pwd = credentials['login_NACIONAL_pwd'] # Contraseña nacional
        self.conn = self.connect_odbc('AS400', maq = 'NACIONAL', user = self.nac_user, pwd = self.nac_pwd)[0] # Conexión ODBC Nacional
        self.conc_date = conc_date # Fecha para el query de conciliación
        self.conc_query = self._get_query_conciliacion() # Dataframe con el query de conciliación

    def _get_query_conciliacion(self):
        """Genera query de conciliación"""
        self.logger.info('Generando query de conciliación')

        # Cargar consulta
        with io.open(os.path.join(self.sql_path, 'query_contabilidad.sql'), mode = 'r', encoding = 'utf8') as f:
            sql = f.read()

        # Reemplazar fecha
        sql = sql.replace('AAAAMMDD', self.conc_date.strftime('%Y%m%d'))

        # Ejecutar query
        df = self.query_AS400(sql, self.conn)[0]

        df.to_excel('./out/conciliación.xlsx', index = False)

        return df
    
    def read_rej_sheet(self):
        """Carga la información que se encuentra en el consolidado del día"""
        # Definir ruta archivo
        self.rej_folder = os.path.join(self.paths['carpeta_rechazos'], self.conc_date.strftime('%Y'), str(self.conc_date.month) + '. ' + self.conc_date.strftime('%B').upper(), self.settings['consolidado_rechazos']['nombre_carpeta'])
        self.consolidation_file = os.path.join(self.rej_folder, self.settings['consolidado_rechazos']['nombre'].replace('AAAAMMDD', self.conc_date.strftime('%Y%m%d')))

        # Cargar hoja de rechazos en Dataframe
        df = pd.read_excel(self.consolidation_file, sheet_name = self.settings['consolidado_rechazos']['hoja_rechazos'], dtype = str)

        # Cambiar formato
        df['Sucursal libros'] = df['Sucursal libros'].astype(int)
        df['Monto'] = df['Monto'].astype(float)

        # Sumar monto y agrupar por sucursal
        self.df_rej = df.groupby(['Sucursal libros'])['Monto'].sum().reset_index()

        return None

    def conciliacion(self):
        """Realiza la conciliación en el consolidado"""
        # Abrir consolidado
        wb = xl.load_workbook(self.consolidation_file)

        # Definir hoja conciliación
        ws_conc = wb[self.settings['consolidado_rechazos']['hoja_conciliacion']]

        # Llenar saldos y diferencia
        for i in range(3, ws_conc.max_row + 1):
            branch = int(ws_conc.cell(row = i, column = 1).value)

            # Llenar saldo nacional
            try:
                ws_conc.cell(row = i, column = 2).value = round(float(self.conc_query[self.conc_query['Oficina'].astype(int) == branch].reset_index().loc[0, 'Saldo']), 2)
            except KeyError:
                ws_conc.cell(row = i, column = 2).value = round(0.0, 2)

            # Llenar oficina y saldo rechazos
            ws_conc.cell(row = i, column = 6).value = branch

            try:
                ws_conc.cell(row = i, column = 7).value = round(float(self.df_rej[self.df_rej['Sucursal libros'] == branch].reset_index().loc[0, 'Monto']), 2)
            except KeyError:
                ws_conc.cell(row = i, column = 7).value = round(0.0, 2)
            
            # Poner fórmula diferencia
            ws_conc["D" + str(i)] = "=B{}-G{}".format(str(i), str(i))

            # Poner en color naranja si la diferencia es mayor a cero
            if ws_conc.cell(row = i, column= 2).value != ws_conc.cell(row = i, column= 7).value:
                ws_conc["D" + str(i)].fill = PatternFill("solid", start_color = "ED7D31")

        # Guardar
        wb.save(self.consolidation_file)

        return None

    def run(self):
        """Orquesta la ejecución de la clase"""
        self.read_rej_sheet()
        self.conciliacion()

        return None

